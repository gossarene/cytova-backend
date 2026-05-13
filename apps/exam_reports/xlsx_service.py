"""
Cytova — XLSX renderer for the Exams-by-Partner pivot.

The rendered workbook mirrors the JSON pivot 1:1 so what the user
sees on screen is exactly what they get in Excel:

  | Exam family | Exam code | Exam name | Partner A | Partner B | ... | Total |
  | Hematology  | NFS       | Numé...   | 133       | 0         | ... | 133   |
  | Hematology  | RETICULO  | ...       | 3         | 0         | ... | 3     |
  | Hematology subtotal     |           | 136       | 0         | ... | 136   |
  | Biochemistry| TP        | ...       | 21        | 10        | ... | 31    |
  | Biochemistry subtotal   |           | 21        | 10        | ... | 31    |
  | Grand total             |           | 157       | 10        | ... | 167   |

Bold styling on header / subtotal / grand-total rows; numeric
cells render with thousands separators so large counts stay
readable. Amount columns appear AFTER the count columns when the
caller opted into ``include_amount`` — keeping the primary metric
(count) as the leftmost partner column band.
"""
from __future__ import annotations

import io
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


HEADER_FONT = Font(name='Calibri', size=11, bold=True, color='FFFFFFFF')
HEADER_FILL = PatternFill('solid', fgColor='FF1F2937')  # slate-800
SUBTOTAL_FONT = Font(name='Calibri', size=11, bold=True)
SUBTOTAL_FILL = PatternFill('solid', fgColor='FFE2E8F0')  # slate-200
GRAND_FONT = Font(name='Calibri', size=11, bold=True, color='FFFFFFFF')
GRAND_FILL = PatternFill('solid', fgColor='FF334155')  # slate-700
CENTER = Alignment(horizontal='center', vertical='center')
LEFT = Alignment(horizontal='left', vertical='center')
RIGHT = Alignment(horizontal='right', vertical='center')
THIN_BORDER = Border(
    left=Side(style='thin', color='FFCBD5E1'),
    right=Side(style='thin', color='FFCBD5E1'),
    top=Side(style='thin', color='FFCBD5E1'),
    bottom=Side(style='thin', color='FFCBD5E1'),
)

# Number formats — counts are plain integers, amounts are 2dp with
# thousand separators. Excel applies the format; we feed it raw
# numerics so column-sum formulas (if a user adds them) still work.
COUNT_FMT = '#,##0;-#,##0;"-"'
AMOUNT_FMT = '#,##0.00;-#,##0.00;"-"'


def render_exams_by_partner_xlsx(report: dict[str, Any]) -> bytes:
    """Return the workbook bytes for the given pivot payload."""
    include_amount = bool(
        report.get('filters_applied', {}).get('include_amount'),
    )
    partners: list[dict[str, str]] = report['partners']
    rows: list[dict[str, Any]] = report['rows']
    subtotals: dict[str, dict[str, Any]] = report['subtotals']
    grand: dict[str, Any] = report['grand_total']

    wb = Workbook()
    ws = wb.active
    ws.title = 'Exams by partner'

    # ---- Header row ----------------------------------------------------
    fixed_cols = ['Exam family', 'Exam code', 'Exam name']
    count_cols = [p['name'] for p in partners] + ['Total']
    if include_amount:
        amount_cols = [f'{p["name"]} (amount)' for p in partners] + ['Total (amount)']
    else:
        amount_cols = []
    header = fixed_cols + count_cols + amount_cols
    ws.append(header)

    for col_idx in range(1, len(header) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER
        cell.border = THIN_BORDER

    # ---- Body — emit rows grouped by family with subtotal after each ----
    # ``rows`` is already sorted by (family_order, family_name, exam_code).
    # We emit each exam row, and when the family key changes from one
    # row to the next we flush the PREVIOUS family's subtotal before
    # continuing. The final family's subtotal is flushed after the
    # loop ends.
    subtotals_by_key = _index_subtotals(subtotals)
    previous_family_key: str | None = None
    for row in rows:
        family_key = row['exam_family_id'] or ''
        if previous_family_key is not None and family_key != previous_family_key:
            sub = subtotals_by_key.get(previous_family_key)
            if sub is not None:
                _append_subtotal_row(ws, sub, partners, include_amount)
        _append_exam_row(ws, row, partners, include_amount)
        previous_family_key = family_key

    if previous_family_key is not None:
        sub = subtotals_by_key.get(previous_family_key)
        if sub is not None:
            _append_subtotal_row(ws, sub, partners, include_amount)

    # ---- Grand total ----------------------------------------------------
    _append_grand_total_row(ws, grand, partners, include_amount)

    # ---- Column widths --------------------------------------------------
    # First three (fixed) columns get a wider default; partner
    # columns auto-scale to the header label length.
    ws.column_dimensions['A'].width = 22
    ws.column_dimensions['B'].width = 14
    ws.column_dimensions['C'].width = 32
    for i, name in enumerate(count_cols):
        col_letter = get_column_letter(3 + 1 + i)
        ws.column_dimensions[col_letter].width = max(12, min(28, len(name) + 2))
    for i, name in enumerate(amount_cols):
        col_letter = get_column_letter(3 + len(count_cols) + 1 + i)
        ws.column_dimensions[col_letter].width = max(14, min(30, len(name) + 2))

    # Freeze the header + the three fixed left columns so a horizontal
    # scroll keeps row identity visible.
    ws.freeze_panes = 'D2'

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ---------------------------------------------------------------------------
# Internal helpers
# ---------------------------------------------------------------------------

def _index_subtotals(subtotals: dict[str, dict[str, Any]]) -> dict[str, dict[str, Any]]:
    """Re-key the subtotal dict by ``family_id`` (stringified, '' for
    None) so the row-emission loop can look it up cheaply."""
    out: dict[str, dict[str, Any]] = {}
    for _key, sub in subtotals.items():
        fid = sub.get('family_id') or ''
        out[str(fid) if fid else ''] = sub
    return out


def _append_exam_row(ws, row, partners, include_amount) -> None:
    values = [
        row['exam_family_name'],
        row['exam_code'],
        row['exam_name'],
    ]
    for p in partners:
        values.append(row['counts'].get(p['id'], 0))
    values.append(row['total'])
    if include_amount:
        amounts = row.get('amounts', {}) or {}
        for p in partners:
            values.append(float(amounts.get(p['id'], 0)))
        values.append(float(row.get('total_amount', 0)))
    ws.append(values)
    _style_data_row(ws, len(partners), include_amount)


def _append_subtotal_row(ws, sub, partners, include_amount) -> None:
    label = f'{sub["family_name"]} subtotal'
    values: list[Any] = [label, '', '']
    for p in partners:
        values.append(sub['counts'].get(p['id'], 0))
    values.append(sub['total'])
    if include_amount:
        amounts = sub.get('amounts', {}) or {}
        for p in partners:
            values.append(float(amounts.get(p['id'], 0)))
        values.append(float(sub.get('total_amount', 0)))
    ws.append(values)
    _style_subtotal_row(ws, len(partners), include_amount)


def _append_grand_total_row(ws, grand, partners, include_amount) -> None:
    values: list[Any] = ['Grand total', '', '']
    for p in partners:
        values.append(grand.get('counts', {}).get(p['id'], 0))
    values.append(grand.get('total', 0))
    if include_amount:
        amounts = grand.get('amounts', {}) or {}
        for p in partners:
            values.append(float(amounts.get(p['id'], 0)))
        values.append(float(grand.get('total_amount', 0)))
    ws.append(values)
    _style_grand_total_row(ws, len(partners), include_amount)


def _style_data_row(ws, partner_count: int, include_amount: bool) -> None:
    row_idx = ws.max_row
    fixed_end = 3
    count_start = fixed_end + 1
    count_end = count_start + partner_count  # +1 for Total
    for col_idx in range(1, count_end + 1):
        cell = ws.cell(row=row_idx, column=col_idx)
        cell.border = THIN_BORDER
        if col_idx <= fixed_end:
            cell.alignment = LEFT
        else:
            cell.alignment = RIGHT
            cell.number_format = COUNT_FMT
    if include_amount:
        amount_start = count_end + 1
        amount_end = amount_start + partner_count
        for col_idx in range(amount_start, amount_end + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.border = THIN_BORDER
            cell.alignment = RIGHT
            cell.number_format = AMOUNT_FMT


def _style_subtotal_row(ws, partner_count: int, include_amount: bool) -> None:
    row_idx = ws.max_row
    _style_data_row(ws, partner_count, include_amount)
    for col_idx in range(1, ws.max_column + 1):
        cell = ws.cell(row=row_idx, column=col_idx)
        cell.font = SUBTOTAL_FONT
        cell.fill = SUBTOTAL_FILL


def _style_grand_total_row(ws, partner_count: int, include_amount: bool) -> None:
    row_idx = ws.max_row
    _style_data_row(ws, partner_count, include_amount)
    for col_idx in range(1, ws.max_column + 1):
        cell = ws.cell(row=row_idx, column=col_idx)
        cell.font = GRAND_FONT
        cell.fill = GRAND_FILL
