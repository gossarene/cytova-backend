"""
Cytova — Exams-by-Partner pivot report tests.

Pin the contract of the service composer + xlsx renderer:

  1. Aggregation correctness — one item per partner / exam yields
     one cell with count=1; multiple items aggregate.
  2. Pivot structure — partners are ordered by name; "Direct"
     synthetic column is last; rows are ordered by family display
     order then exam code.
  3. Subtotals and grand total are arithmetically consistent.
  4. Filters — date range, partner_ids, family_ids, status filters
     all narrow the result set correctly.
  5. Tenant isolation — items from a different tenant schema do
     NOT bleed into the report. Verified through the schema_context
     boundary.
  6. Empty dataset returns a valid empty payload, not a 500.
  7. XLSX export — produces a non-empty workbook that openpyxl
     can re-parse and whose first sheet carries the expected header
     row + a grand-total row.
"""
from __future__ import annotations

import io
from datetime import date, timedelta
from decimal import Decimal

import pytest
from django.utils import timezone
from django_tenants.utils import get_public_schema_name, schema_context
from openpyxl import load_workbook

from apps.catalog.models import (
    ExamCategory, ExamDefinition, ExamFamily, ExamTechnique,
    ResultStructure, SampleType,
)
from apps.exam_reports.services import (
    DIRECT_PARTNER_KEY, DIRECT_PARTNER_LABEL,
    EXAM_PROGRESS_ALL, EXAM_PROGRESS_IN_PROGRESS,
    EXAM_PROGRESS_PERFORMED, EXAM_PROGRESS_REJECTED,
    ExamsByPartnerFilters, build_exams_by_partner_report,
)
from apps.exam_reports.xlsx_service import render_exams_by_partner_xlsx
from apps.partners.models import PartnerOrganization
from apps.patients.models import Patient
from apps.requests.models import (
    AnalysisRequest, AnalysisRequestItem, RequestStatus, SourceType,
)


# ---------------------------------------------------------------------------
# Subscription fixture (matches the convention used in the requests tests)
# ---------------------------------------------------------------------------

@pytest.fixture(autouse=True)
def _usable_subscription(_test_tenant_schema, django_db_blocker):
    from apps.tenants.models import (
        Subscription, SubscriptionPlan, SubscriptionStatus, Tenant,
    )
    with django_db_blocker.unblock():
        with schema_context(get_public_schema_name()):
            plan, _ = SubscriptionPlan.objects.get_or_create(
                code='TEST_TRIAL',
                defaults={
                    'name': 'Test Trial', 'is_trial': True,
                    'trial_duration_days': 30, 'is_public': False,
                },
            )
            tenant = Tenant.objects.get(schema_name=_test_tenant_schema)
            Subscription.objects.get_or_create(
                tenant=tenant, status=SubscriptionStatus.TRIAL,
                defaults={'plan': plan},
            )
    yield


# ---------------------------------------------------------------------------
# Domain fixtures
# ---------------------------------------------------------------------------

@pytest.fixture()
def category():
    return ExamCategory.objects.create(name='Default', display_order=1)


@pytest.fixture()
def hematology_family():
    return ExamFamily.objects.create(name='Hematology', display_order=1)


@pytest.fixture()
def biochemistry_family():
    return ExamFamily.objects.create(name='Biochemistry', display_order=2)


@pytest.fixture()
def technique():
    return ExamTechnique.objects.create(name='Default')


@pytest.fixture()
def nfs(hematology_family, technique, category):
    return ExamDefinition.objects.create(
        category=category, family=hematology_family, technique=technique,
        code='NFS', name='Numération Formule Sanguine',
        sample_type=SampleType.BLOOD,
        result_structure=ResultStructure.SINGLE_VALUE,
        unit_price=Decimal('20'),
    )


@pytest.fixture()
def reticulo(hematology_family, technique, category):
    return ExamDefinition.objects.create(
        category=category, family=hematology_family, technique=technique,
        code='RETICULO', name='Reticulocytes',
        sample_type=SampleType.BLOOD,
        result_structure=ResultStructure.SINGLE_VALUE,
        unit_price=Decimal('15'),
    )


@pytest.fixture()
def tp(biochemistry_family, technique, category):
    return ExamDefinition.objects.create(
        category=category, family=biochemistry_family, technique=technique,
        code='TP', name='Total Protein',
        sample_type=SampleType.BLOOD,
        result_structure=ResultStructure.SINGLE_VALUE,
        unit_price=Decimal('10'),
    )


@pytest.fixture()
def partner_a():
    return PartnerOrganization.objects.create(name='SERENA', code='SERENA')


@pytest.fixture()
def partner_b():
    return PartnerOrganization.objects.create(name='Biosso', code='BIOSSO')


@pytest.fixture()
def patient(lab_admin):
    return Patient.objects.create(
        document_type='NATIONAL_ID_CARD',
        document_number='NID-EXAM-001',
        first_name='Eva', last_name='Stat',
        date_of_birth=date(1990, 1, 1), gender='FEMALE',
        created_by=lab_admin,
    )


import uuid


def _make_request(
    *,
    patient, exams, partner=None, status=RequestStatus.VALIDATED,
    confirmed_at=None, lab_admin,
    item_status='VALIDATED',
    item_execution_mode='INTERNAL',
) -> AnalysisRequest:
    """Insert a finished-state request + its items directly in the
    DB, bypassing the service to keep the test focused on the
    composer's behaviour (the service is exercised by other tests).

    Items default to ``status='VALIDATED'`` so they fall into the
    PERFORMED exam-progress group — that matches what the existing
    test scenarios are asserting ("an exam was performed"). Specific
    tests override ``item_status`` / ``item_execution_mode`` to seed
    IN_PROGRESS or REJECTED rows.

    ``request_number`` + ``public_reference`` are unique columns
    that the production service populates after first save; here we
    seed unique-per-call values so multiple inserts in a single
    test don't collide on the unique constraint."""
    suffix = uuid.uuid4().hex[:8].upper()
    ar = AnalysisRequest.objects.create(
        patient=patient, created_by=lab_admin,
        source_type=(
            SourceType.PARTNER_ORGANIZATION if partner
            else SourceType.DIRECT_PATIENT
        ),
        partner_organization=partner,
        status=status,
        confirmed_at=confirmed_at or timezone.now(),
        request_number=f'TEST-REQ-{suffix}',
        public_reference=f'TEST-{suffix}',
    )
    for exam in exams:
        AnalysisRequestItem.objects.create(
            analysis_request=ar,
            exam_definition=exam,
            unit_price=exam.unit_price,
            billed_price=exam.unit_price,
            result_structure_snapshot=exam.result_structure,
            status=item_status,
            execution_mode=item_execution_mode,
        )
    return ar


def _filters(**overrides) -> ExamsByPartnerFilters:
    today = date.today()
    base = dict(
        period_start=today - timedelta(days=30),
        period_end=today,
    )
    base.update(overrides)
    return ExamsByPartnerFilters(**base)


# ===========================================================================
# 1. Aggregation + pivot structure
# ===========================================================================

@pytest.mark.django_db
class TestAggregation:

    def test_single_item_one_cell(
        self, patient, lab_admin, nfs, partner_a,
    ):
        _make_request(
            patient=patient, exams=[nfs], partner=partner_a,
            lab_admin=lab_admin,
        )
        report = build_exams_by_partner_report(_filters())
        assert len(report['rows']) == 1
        row = report['rows'][0]
        assert row['exam_code'] == 'NFS'
        assert row['exam_family_name'] == 'Hematology'
        assert row['counts'] == {str(partner_a.id): 1}
        assert row['total'] == 1
        assert report['partners'] == [{'id': str(partner_a.id), 'name': 'SERENA'}]
        assert report['grand_total']['total'] == 1

    def test_multiple_partners_multiple_exams(
        self, patient, lab_admin, nfs, reticulo, tp,
        partner_a, partner_b,
    ):
        # SERENA: 3 × NFS, 1 × TP
        for _ in range(3):
            _make_request(
                patient=patient, exams=[nfs], partner=partner_a,
                lab_admin=lab_admin,
            )
        _make_request(
            patient=patient, exams=[tp], partner=partner_a,
            lab_admin=lab_admin,
        )
        # Biosso: 2 × TP, 1 × RETICULO
        for _ in range(2):
            _make_request(
                patient=patient, exams=[tp], partner=partner_b,
                lab_admin=lab_admin,
            )
        _make_request(
            patient=patient, exams=[reticulo], partner=partner_b,
            lab_admin=lab_admin,
        )

        report = build_exams_by_partner_report(_filters())

        # Partners sorted by name (case-insensitive): Biosso, SERENA.
        assert [p['name'] for p in report['partners']] == ['Biosso', 'SERENA']

        # Hematology before Biochemistry by display_order.
        family_order = [r['exam_family_name'] for r in report['rows']]
        # Within Hematology, exams sorted by code ascending.
        assert family_order[:2] == ['Hematology', 'Hematology']
        assert [r['exam_code'] for r in report['rows'][:2]] == ['NFS', 'RETICULO']

        # Counts pivot — NFS only from SERENA.
        nfs_row = next(r for r in report['rows'] if r['exam_code'] == 'NFS')
        assert nfs_row['counts'].get(str(partner_a.id)) == 3
        assert nfs_row['counts'].get(str(partner_b.id)) is None
        assert nfs_row['total'] == 3

        tp_row = next(r for r in report['rows'] if r['exam_code'] == 'TP')
        assert tp_row['counts'].get(str(partner_a.id)) == 1
        assert tp_row['counts'].get(str(partner_b.id)) == 2
        assert tp_row['total'] == 3

        # Subtotals per family + grand total.
        hem_sub = next(
            s for s in report['subtotals'].values()
            if s['family_name'] == 'Hematology'
        )
        bio_sub = next(
            s for s in report['subtotals'].values()
            if s['family_name'] == 'Biochemistry'
        )
        assert hem_sub['total'] == 4  # 3 NFS + 1 RETICULO
        assert bio_sub['total'] == 3  # 1 + 2 TP
        assert report['grand_total']['total'] == 7
        assert report['grand_total']['counts'] == {
            str(partner_a.id): 4,
            str(partner_b.id): 3,
        }

    def test_direct_patient_appears_as_synthetic_column(
        self, patient, lab_admin, nfs,
    ):
        _make_request(
            patient=patient, exams=[nfs], partner=None,
            lab_admin=lab_admin,
        )
        report = build_exams_by_partner_report(_filters())
        assert report['partners'] == [
            {'id': DIRECT_PARTNER_KEY, 'name': DIRECT_PARTNER_LABEL},
        ]
        assert report['rows'][0]['counts'] == {DIRECT_PARTNER_KEY: 1}

    def test_direct_column_renders_last_when_partners_exist(
        self, patient, lab_admin, nfs, partner_a,
    ):
        _make_request(
            patient=patient, exams=[nfs], partner=partner_a,
            lab_admin=lab_admin,
        )
        _make_request(
            patient=patient, exams=[nfs], partner=None,
            lab_admin=lab_admin,
        )
        report = build_exams_by_partner_report(_filters())
        assert [p['id'] for p in report['partners']] == [
            str(partner_a.id), DIRECT_PARTNER_KEY,
        ]


# ===========================================================================
# 2. Filters
# ===========================================================================

@pytest.mark.django_db
class TestFilters:

    def test_date_range_filter_excludes_outside_period(
        self, patient, lab_admin, nfs, partner_a,
    ):
        # Inside window (yesterday)
        _make_request(
            patient=patient, exams=[nfs], partner=partner_a,
            confirmed_at=timezone.now() - timedelta(days=1),
            lab_admin=lab_admin,
        )
        # Outside window (40 days ago)
        _make_request(
            patient=patient, exams=[nfs], partner=partner_a,
            confirmed_at=timezone.now() - timedelta(days=40),
            lab_admin=lab_admin,
        )
        today = date.today()
        report = build_exams_by_partner_report(_filters(
            period_start=today - timedelta(days=7),
            period_end=today,
        ))
        assert report['grand_total']['total'] == 1

    def test_partner_filter_narrows_to_selected_partners(
        self, patient, lab_admin, nfs, partner_a, partner_b,
    ):
        _make_request(
            patient=patient, exams=[nfs], partner=partner_a,
            lab_admin=lab_admin,
        )
        _make_request(
            patient=patient, exams=[nfs], partner=partner_b,
            lab_admin=lab_admin,
        )
        report = build_exams_by_partner_report(_filters(
            partner_ids=(str(partner_a.id),),
            include_direct=False,
        ))
        assert {p['id'] for p in report['partners']} == {str(partner_a.id)}
        assert report['grand_total']['total'] == 1

    def test_family_filter_narrows_rows(
        self, patient, lab_admin, nfs, tp, partner_a,
    ):
        _make_request(
            patient=patient, exams=[nfs, tp], partner=partner_a,
            lab_admin=lab_admin,
        )
        report = build_exams_by_partner_report(_filters(
            exam_family_ids=(str(nfs.family_id),),
        ))
        assert all(r['exam_family_name'] == 'Hematology' for r in report['rows'])
        assert report['grand_total']['total'] == 1

    def test_request_status_filter_default_excludes_in_progress(
        self, patient, lab_admin, nfs, partner_a,
    ):
        # An IN_ANALYSIS request is NOT "performed" by the default
        # request-status filter and must be excluded.
        _make_request(
            patient=patient, exams=[nfs], partner=partner_a,
            status=RequestStatus.IN_ANALYSIS,
            lab_admin=lab_admin,
        )
        report = build_exams_by_partner_report(_filters())
        assert report['grand_total']['total'] == 0

    def test_rejected_items_excluded(
        self, patient, lab_admin, nfs, partner_a,
    ):
        ar = _make_request(
            patient=patient, exams=[nfs], partner=partner_a,
            lab_admin=lab_admin,
        )
        ar.items.update(execution_mode='REJECTED')
        report = build_exams_by_partner_report(_filters())
        assert report['grand_total']['total'] == 0


# ===========================================================================
# 3. Amount addon
# ===========================================================================

@pytest.mark.django_db
class TestAmountColumns:

    def test_include_amount_adds_monetary_totals(
        self, patient, lab_admin, nfs, partner_a,
    ):
        # Two NFS for SERENA at 20.00 each — expect amount = 40.00.
        for _ in range(2):
            _make_request(
                patient=patient, exams=[nfs], partner=partner_a,
                lab_admin=lab_admin,
            )
        report = build_exams_by_partner_report(_filters(include_amount=True))
        row = report['rows'][0]
        # ``amounts`` is keyed by partner id; serialised as 2-dp string.
        assert row['amounts'][str(partner_a.id)] == '40.00'
        assert row['total_amount'] == '40.00'
        assert report['grand_total']['amounts'][str(partner_a.id)] == '40.00'
        assert report['grand_total']['total_amount'] == '40.00'

    def test_default_omits_amount_columns(
        self, patient, lab_admin, nfs, partner_a,
    ):
        _make_request(
            patient=patient, exams=[nfs], partner=partner_a,
            lab_admin=lab_admin,
        )
        report = build_exams_by_partner_report(_filters())
        # Rows must NOT carry 'amounts' / 'total_amount' when the
        # primary metric (count) is the only thing requested.
        assert 'amounts' not in report['rows'][0]
        assert 'total_amount' not in report['rows'][0]


# ===========================================================================
# 4. Exam-progress filter
# ===========================================================================
#
# These tests exercise the item-level progress grouping that drives
# the "Exam status" filter on the UI. Every test seeds a mix of
# item states so the per-group counters can't shortcut a partial
# match.

@pytest.mark.django_db
class TestExamProgressFilter:

    def test_performed_counts_validated_and_completed_only(
        self, patient, lab_admin, nfs, partner_a,
    ):
        _make_request(
            patient=patient, exams=[nfs], partner=partner_a,
            lab_admin=lab_admin, item_status='VALIDATED',
        )
        _make_request(
            patient=patient, exams=[nfs], partner=partner_a,
            lab_admin=lab_admin, item_status='COMPLETED',
        )
        # Still in analysis — NOT performed.
        _make_request(
            patient=patient, exams=[nfs], partner=partner_a,
            lab_admin=lab_admin,
            status=RequestStatus.IN_ANALYSIS,
            item_status='UNDER_REVIEW',
        )
        # Rejected — NOT performed.
        _make_request(
            patient=patient, exams=[nfs], partner=partner_a,
            lab_admin=lab_admin, item_status='REJECTED',
        )

        report = build_exams_by_partner_report(_filters(
            exam_progress_status=EXAM_PROGRESS_PERFORMED,
        ))
        assert report['grand_total']['total'] == 2

    def test_in_progress_counts_workflow_states(
        self, patient, lab_admin, nfs, partner_a,
    ):
        _make_request(
            patient=patient, exams=[nfs], partner=partner_a,
            lab_admin=lab_admin,
            status=RequestStatus.IN_ANALYSIS,
            item_status='UNDER_REVIEW',
        )
        _make_request(
            patient=patient, exams=[nfs], partner=partner_a,
            lab_admin=lab_admin,
            status=RequestStatus.COLLECTION_IN_PROGRESS,
            item_status='COLLECTED',
        )
        # Performed — NOT in progress.
        _make_request(
            patient=patient, exams=[nfs], partner=partner_a,
            lab_admin=lab_admin, item_status='VALIDATED',
        )
        # Rejected — NOT in progress.
        _make_request(
            patient=patient, exams=[nfs], partner=partner_a,
            lab_admin=lab_admin, item_status='REJECTED',
        )

        # IN_PROGRESS group with no explicit request_statuses — the
        # serializer-side default suppression is exercised through
        # the dataclass: an empty request_statuses tuple means "no
        # parent-status filter".
        report = build_exams_by_partner_report(_filters(
            exam_progress_status=EXAM_PROGRESS_IN_PROGRESS,
            request_statuses=(),
        ))
        assert report['grand_total']['total'] == 2

    def test_rejected_counts_both_signals(
        self, patient, lab_admin, nfs, partner_a,
    ):
        # ItemStatus.REJECTED.
        _make_request(
            patient=patient, exams=[nfs], partner=partner_a,
            lab_admin=lab_admin, item_status='REJECTED',
        )
        # ExecutionMode.REJECTED with a non-REJECTED workflow status.
        _make_request(
            patient=patient, exams=[nfs], partner=partner_a,
            lab_admin=lab_admin,
            item_status='VALIDATED',  # ignored for REJECTED group
            item_execution_mode='REJECTED',
        )
        # Performed — NOT rejected.
        _make_request(
            patient=patient, exams=[nfs], partner=partner_a,
            lab_admin=lab_admin, item_status='VALIDATED',
        )

        report = build_exams_by_partner_report(_filters(
            exam_progress_status=EXAM_PROGRESS_REJECTED,
            request_statuses=(),
        ))
        assert report['grand_total']['total'] == 2

    def test_all_counts_every_state(
        self, patient, lab_admin, nfs, partner_a,
    ):
        # 1 performed + 1 in-progress + 1 rejected.
        _make_request(
            patient=patient, exams=[nfs], partner=partner_a,
            lab_admin=lab_admin, item_status='VALIDATED',
        )
        _make_request(
            patient=patient, exams=[nfs], partner=partner_a,
            lab_admin=lab_admin,
            status=RequestStatus.IN_ANALYSIS,
            item_status='UNDER_REVIEW',
        )
        _make_request(
            patient=patient, exams=[nfs], partner=partner_a,
            lab_admin=lab_admin, item_status='REJECTED',
        )

        report = build_exams_by_partner_report(_filters(
            exam_progress_status=EXAM_PROGRESS_ALL,
            request_statuses=(),
        ))
        assert report['grand_total']['total'] == 3

    def test_mixed_status_request_counts_per_item(
        self, patient, lab_admin, nfs, reticulo, tp, partner_a,
    ):
        # A SINGLE request with three items in three different
        # states. Each item is counted independently — the parent
        # request's status is irrelevant to the per-item grouping.
        suffix = uuid.uuid4().hex[:8].upper()
        ar = AnalysisRequest.objects.create(
            patient=patient, created_by=lab_admin,
            source_type=SourceType.PARTNER_ORGANIZATION,
            partner_organization=partner_a,
            status=RequestStatus.IN_ANALYSIS,
            confirmed_at=timezone.now(),
            request_number=f'TEST-REQ-{suffix}',
            public_reference=f'TEST-{suffix}',
        )
        AnalysisRequestItem.objects.create(
            analysis_request=ar, exam_definition=nfs,
            unit_price=nfs.unit_price, billed_price=nfs.unit_price,
            result_structure_snapshot=nfs.result_structure,
            status='VALIDATED', execution_mode='INTERNAL',
        )
        AnalysisRequestItem.objects.create(
            analysis_request=ar, exam_definition=reticulo,
            unit_price=reticulo.unit_price, billed_price=reticulo.unit_price,
            result_structure_snapshot=reticulo.result_structure,
            status='UNDER_REVIEW', execution_mode='INTERNAL',
        )
        AnalysisRequestItem.objects.create(
            analysis_request=ar, exam_definition=tp,
            unit_price=tp.unit_price, billed_price=tp.unit_price,
            result_structure_snapshot=tp.result_structure,
            status='REJECTED', execution_mode='INTERNAL',
        )

        # PERFORMED → 1 (NFS only). The parent request is
        # IN_ANALYSIS — the test deliberately suppresses the
        # default parent-status filter to focus on the item-level
        # grouping. The next test
        # (``test_filter_combines_with_request_status``) covers the
        # opposite case where the parent filter is honoured.
        assert build_exams_by_partner_report(_filters(
            exam_progress_status=EXAM_PROGRESS_PERFORMED,
            request_statuses=(),
        ))['grand_total']['total'] == 1

        # IN_PROGRESS → 1 (RETICULO only). Need to clear the
        # default request_status filter because the parent request
        # is IN_ANALYSIS.
        assert build_exams_by_partner_report(_filters(
            exam_progress_status=EXAM_PROGRESS_IN_PROGRESS,
            request_statuses=(),
        ))['grand_total']['total'] == 1

        # REJECTED → 1 (TP only).
        assert build_exams_by_partner_report(_filters(
            exam_progress_status=EXAM_PROGRESS_REJECTED,
            request_statuses=(),
        ))['grand_total']['total'] == 1

        # ALL → 3.
        assert build_exams_by_partner_report(_filters(
            exam_progress_status=EXAM_PROGRESS_ALL,
            request_statuses=(),
        ))['grand_total']['total'] == 3

    def test_filter_combines_with_request_status(
        self, patient, lab_admin, nfs, partner_a,
    ):
        # Validated item under an IN_ANALYSIS parent (e.g. one
        # item finished, others still in progress on the same
        # request). PERFORMED+request_statuses=(IN_ANALYSIS,) should
        # surface it; PERFORMED+request_statuses=(VALIDATED,) should
        # not (parent isn't VALIDATED).
        _make_request(
            patient=patient, exams=[nfs], partner=partner_a,
            lab_admin=lab_admin,
            status=RequestStatus.IN_ANALYSIS,
            item_status='VALIDATED',
        )
        assert build_exams_by_partner_report(_filters(
            exam_progress_status=EXAM_PROGRESS_PERFORMED,
            request_statuses=('IN_ANALYSIS',),
        ))['grand_total']['total'] == 1
        assert build_exams_by_partner_report(_filters(
            exam_progress_status=EXAM_PROGRESS_PERFORMED,
            request_statuses=('VALIDATED',),
        ))['grand_total']['total'] == 0

    def test_serializer_drops_default_request_statuses_for_non_performed(self):
        # When the caller omits ``request_statuses`` AND picks a
        # non-PERFORMED progress group, the serializer must not
        # apply the legacy ``(VALIDATED, COMPLETED, RESULT_ISSUED)``
        # default — otherwise the IN_PROGRESS column would be
        # silently zero on every report.
        from apps.exam_reports.serializers import ExamsByPartnerFiltersSerializer

        ser = ExamsByPartnerFiltersSerializer(data={
            'period_start': '2026-01-01',
            'period_end': '2026-12-31',
            'exam_progress_status': EXAM_PROGRESS_IN_PROGRESS,
        })
        assert ser.is_valid(), ser.errors
        resolved = ser.to_filters()
        assert resolved.exam_progress_status == EXAM_PROGRESS_IN_PROGRESS
        assert resolved.request_statuses == ()

    def test_serializer_keeps_default_for_performed(self):
        from apps.exam_reports.serializers import ExamsByPartnerFiltersSerializer

        ser = ExamsByPartnerFiltersSerializer(data={
            'period_start': '2026-01-01',
            'period_end': '2026-12-31',
            # No exam_progress_status → defaults to PERFORMED.
        })
        assert ser.is_valid(), ser.errors
        resolved = ser.to_filters()
        assert resolved.exam_progress_status == EXAM_PROGRESS_PERFORMED
        # Legacy default preserved.
        assert resolved.request_statuses == (
            'VALIDATED', 'COMPLETED', 'RESULT_ISSUED',
        )

    def test_xlsx_export_respects_progress_filter(
        self, patient, lab_admin, nfs, partner_a,
    ):
        # Two items: one VALIDATED, one REJECTED. The XLSX produced
        # under PERFORMED must NOT contain a row for the rejected
        # item — equivalently, the grand total in the bottom row is
        # 1, not 2.
        _make_request(
            patient=patient, exams=[nfs], partner=partner_a,
            lab_admin=lab_admin, item_status='VALIDATED',
        )
        _make_request(
            patient=patient, exams=[nfs], partner=partner_a,
            lab_admin=lab_admin, item_status='REJECTED',
        )

        report = build_exams_by_partner_report(_filters(
            exam_progress_status=EXAM_PROGRESS_PERFORMED,
        ))
        xlsx_bytes = render_exams_by_partner_xlsx(report)
        wb = load_workbook(io.BytesIO(xlsx_bytes))
        ws = wb.active
        last_row = [c.value for c in ws[ws.max_row]]
        assert last_row[0] == 'Grand total'
        assert last_row[-1] == 1

        # Switching the filter to REJECTED gives the opposite count.
        report_rej = build_exams_by_partner_report(_filters(
            exam_progress_status=EXAM_PROGRESS_REJECTED,
            request_statuses=(),
        ))
        xlsx_rej = render_exams_by_partner_xlsx(report_rej)
        wb_rej = load_workbook(io.BytesIO(xlsx_rej))
        last_row_rej = [c.value for c in wb_rej.active[wb_rej.active.max_row]]
        assert last_row_rej[-1] == 1


# ===========================================================================
# 5. Empty dataset
# ===========================================================================

@pytest.mark.django_db
class TestEmpty:

    def test_no_data_returns_well_formed_empty_payload(self):
        report = build_exams_by_partner_report(_filters())
        assert report['partners'] == []
        assert report['rows'] == []
        assert report['subtotals'] == {}
        assert report['grand_total'] == {'counts': {}, 'total': 0}


# ===========================================================================
# 5. Tenant isolation (smoke test)
# ===========================================================================

@pytest.mark.django_db
class TestTenantIsolation:

    def test_report_query_runs_in_active_tenant_schema(
        self, patient, lab_admin, nfs, partner_a,
    ):
        # Sanity inside the active tenant: the row inserted here is
        # found by the composer. The composer issues a single
        # AnalysisRequestItem query — by design no explicit
        # ``tenant_id`` filter is needed because django-tenants
        # swaps the ``search_path`` per request. We pin that
        # behaviour by switching to the ``public`` schema and
        # confirming the composer raises rather than silently
        # surfacing tenant rows: the tenant tables are physically
        # not visible from ``public``, which is the load-bearing
        # multi-tenant safety property.
        _make_request(
            patient=patient, exams=[nfs], partner=partner_a,
            lab_admin=lab_admin,
        )
        assert build_exams_by_partner_report(
            _filters(),
        )['grand_total']['total'] == 1

        from django.db.utils import ProgrammingError
        with schema_context(get_public_schema_name()):
            with pytest.raises(ProgrammingError):
                build_exams_by_partner_report(_filters())


# ===========================================================================
# 6. XLSX export
# ===========================================================================

@pytest.mark.django_db
class TestXlsxExport:

    def test_xlsx_is_valid_and_carries_grand_total(
        self, patient, lab_admin, nfs, tp, partner_a, partner_b,
    ):
        for _ in range(3):
            _make_request(
                patient=patient, exams=[nfs], partner=partner_a,
                lab_admin=lab_admin,
            )
        _make_request(
            patient=patient, exams=[tp], partner=partner_b,
            lab_admin=lab_admin,
        )
        report = build_exams_by_partner_report(_filters())
        xlsx_bytes = render_exams_by_partner_xlsx(report)
        assert xlsx_bytes  # non-empty

        wb = load_workbook(io.BytesIO(xlsx_bytes))
        ws = wb.active
        # Header row: fixed columns + partner names + Total.
        header = [c.value for c in ws[1]]
        assert header[:3] == ['Exam family', 'Exam code', 'Exam name']
        assert 'SERENA' in header
        assert 'Biosso' in header
        assert header[-1] == 'Total'

        # Grand total row exists at the bottom with the right total.
        last_row = [c.value for c in ws[ws.max_row]]
        assert last_row[0] == 'Grand total'
        assert last_row[-1] == 4  # 3 NFS + 1 TP

    def test_xlsx_empty_dataset_still_generates(self):
        report = build_exams_by_partner_report(_filters())
        xlsx_bytes = render_exams_by_partner_xlsx(report)
        wb = load_workbook(io.BytesIO(xlsx_bytes))
        ws = wb.active
        # Header is always present; grand total row appears even
        # when zero data rows exist.
        assert ws[1][0].value == 'Exam family'
        last_row = [c.value for c in ws[ws.max_row]]
        assert last_row[0] == 'Grand total'
